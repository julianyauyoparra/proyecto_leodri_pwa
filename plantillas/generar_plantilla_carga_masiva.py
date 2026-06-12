#!/usr/bin/env python3
"""Genera la plantilla Excel para carga masiva de productos LEODRI."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path(__file__).resolve().parent / "LEODRI_carga_masiva_productos.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
REQUIRED_FILL = PatternFill("solid", fgColor="FFF3CD")
NOTE_FONT = Font(size=10, color="444444")


def style_header(ws, headers):
    for col, (title, width, required) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
        if required:
            ws.cell(row=2, column=col).fill = REQUIRED_FILL
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"


def add_instructions(wb):
    ws = wb.create_sheet("INSTRUCCIONES", 0)
    ws.column_dimensions["A"].width = 100
    lines = [
        "PLANTILLA DE CARGA MASIVA — LEODRI",
        "",
        "1. codigo_referencia: identificador único por producto (ej. PROD-001). Enlaza todas las hojas.",
        "2. Complete PRODUCTOS (1 fila por zapato), luego COLORES, TALLAS, STOCK y BENEFICIOS.",
        "3. bullets y tags: valores separados por | (pipe). Ej: Luces LED|Suela antideslizante",
        "4. img_derecha es obligatoria (miniatura del catálogo). Rutas relativas: assets/productos/...",
        "5. sku_base usa {talla} como placeholder. Ej: KDF-SPL-RJN-LED-{talla}",
        "6. STOCK alimenta inventario_variantes (Railway). stock=0 deja la talla sin disponibilidad.",
        "7. icono beneficios: check | zap | shoe | wave | shield",
        "8. Las filas de ejemplo (PROD-EJEMPLO-001) pueden borrarse antes de importar.",
        "",
        "Tablas destino: productos, producto_colores, producto_tallas, producto_beneficios, inventario_variantes",
    ]
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = NOTE_FONT if i > 1 else Font(bold=True, size=14)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def sheet_productos(wb):
    ws = wb.create_sheet("PRODUCTOS")
    headers = [
        ("codigo_referencia*", 18, True),
        ("marca*", 14, True),
        ("nombre*", 32, True),
        ("descripcion", 48, False),
        ("bullets", 40, False),
        ("tags", 24, False),
        ("precio*", 10, True),
        ("color_default", 14, False),
        ("activo", 8, False),
        ("orden", 8, False),
    ]
    style_header(ws, headers)
    ws.append([
        "PROD-EJEMPLO-001",
        "KidsFoot",
        "Zapatillas Niño Spider Light",
        "Haz que sus aventuras diarias sean más divertidas y cómodas con estas zapatillas de diseño heroico.",
        "Luces LED|Suela antideslizante|Diseño Spider",
        "Niño|Running",
        89.9,
        "RJN",
        1,
        1,
    ])
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"I2:I5000")


def sheet_colores(wb):
    ws = wb.create_sheet("COLORES")
    headers = [
        ("codigo_referencia*", 18, True),
        ("codigo_color*", 14, True),
        ("etiqueta*", 22, True),
        ("codigo_inventario", 18, False),
        ("sku_base", 28, False),
        ("alt", 36, False),
        ("img_frente", 36, False),
        ("img_derecha*", 36, True),
        ("img_izquierda", 36, False),
        ("img_posterior", 36, False),
        ("img_arriba", 36, False),
        ("img_abajo", 36, False),
        ("orden", 8, False),
    ]
    style_header(ws, headers)
    rows = [
        [
            "PROD-EJEMPLO-001", "RJN", "rojo y negro", "KDF-SPL-RJN-LED",
            "KDF-SPL-RJN-LED-{talla}", "KidsFoot Spider Light rojo y negro",
            "assets/demo/hero-rjn.png", "assets/demo/hero-rjn.png", "", "", "", "", 0,
        ],
        [
            "PROD-EJEMPLO-001", "GRY", "gris y negro", "KDF-SPL-GRY-LED",
            "KDF-SPL-GRY-LED-{talla}", "KidsFoot Spider Light gris y negro",
            "assets/demo/color-gry.png", "assets/demo/color-gry.png", "", "", "", "", 1,
        ],
    ]
    for row in rows:
        ws.append(row)


def sheet_tallas(wb):
    ws = wb.create_sheet("TALLAS")
    headers = [
        ("codigo_referencia*", 18, True),
        ("talla_numero*", 12, True),
        ("disponible", 12, False),
        ("orden", 8, False),
    ]
    style_header(ws, headers)
    for i, talla in enumerate(["21", "22", "23", "24", "25", "26"]):
        ws.append(["PROD-EJEMPLO-001", talla, 1 if talla not in ("23", "25") else 0, i])


def sheet_stock(wb):
    ws = wb.create_sheet("STOCK")
    headers = [
        ("codigo_referencia*", 18, True),
        ("codigo_color*", 14, True),
        ("talla_numero*", 12, True),
        ("stock*", 10, True),
        ("sku", 28, False),
    ]
    style_header(ws, headers)
    ws.append(["PROD-EJEMPLO-001", "RJN", "21", 5, ""])
    ws.append(["PROD-EJEMPLO-001", "RJN", "22", 3, ""])
    ws.append(["PROD-EJEMPLO-001", "RJN", "23", 0, ""])
    ws.append(["PROD-EJEMPLO-001", "GRY", "22", 2, "KDF-SPL-GRY-LED-22"])


def sheet_beneficios(wb):
    ws = wb.create_sheet("BENEFICIOS")
    headers = [
        ("codigo_referencia*", 18, True),
        ("icono", 10, False),
        ("titulo*", 32, True),
        ("texto", 56, False),
        ("orden", 8, False),
    ]
    style_header(ws, headers)
    ws.append([
        "PROD-EJEMPLO-001", "zap", "Luces LED de Larga Duración",
        "<strong>Diversión asegurada:</strong> Suelas con luces intermitentes al caminar.", 0,
    ])
    ws.append([
        "PROD-EJEMPLO-001", "check", "Cierre Autónomo (Velcro + Elástico)",
        "<strong>Fácil de poner:</strong> El niño se las pone solo en segundos.", 1,
    ])
    dv = DataValidation(
        type="list",
        formula1='"check,zap,shoe,wave,shield"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add("B2:B5000")


def main():
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb)
    sheet_productos(wb)
    sheet_colores(wb)
    sheet_tallas(wb)
    sheet_stock(wb)
    sheet_beneficios(wb)
    wb.save(OUTPUT)
    print(f"Plantilla generada: {OUTPUT}")


if __name__ == "__main__":
    main()
